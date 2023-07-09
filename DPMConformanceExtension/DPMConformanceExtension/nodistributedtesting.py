from tkinter.filedialog import SaveFileDialog
from traceback import print_tb
from pm4py.objects.log.importer.xes import importer as xes_importer
from DPMConformanceExtension import to_petri_net as pt_converter
from pm4py.algo.conformance.alignments.petri_net import algorithm as at
from DPMConformanceExtension import ptml
import pm4py
import openpyxl as op
import time
from DPMConformanceExtension.conformance_algorithm import call_for_alignment
from DPMConformanceExtension.decompose_conformance_method.version2 import NodePriorityVariants
from DPMConformanceExtension.decompose_conformance_method.version2 import TraceAllocateVariants
from DPMConformanceExtension.decompose_conformance_method.version2 import ConformanceVariants
from DPMConformanceExtension.decompose_conformance_method.version2 import decom_conformance
from DPMConformanceExtension.decompose_conformance_method import version2

event_logs_num = ['prBm6']

for r_ns in range(1):
    for eln in event_logs_num:
        
        print("log",eln)
        if eln in [1,2,3,4,5]:
            log_path = "/home/hadoop/Projects/testdata/data/L"+str(eln)+"/L"+str(eln)+".xes"
            model_path = "/home/hadoop/Projects/testdata/data/L"+str(eln)+"/L"+str(eln)+".ptml"

            r_path = r"/home/hadoop/Projects/exresult/nnactnum_tarset.xlsx"
            s_path = "/home/hadoop/Projects/exresult/nnactnum_tarset.xlsx"
        else:
            log_path = "/home/hadoop/Projects/testdata/data/BPM2013/newlog/new_"+eln+".xes"
            model_path = "/home/hadoop/Projects/testdata/data/BPM2013/Multitree/"+eln+".ptml"

            r_path = r"/home/hadoop/Projects/exresult/Annactnum_tarset.xlsx"
            s_path = "/home/hadoop/Projects/exresult/Annactnum_tarset.xlsx"
            o_path = "/home/hadoop/Projects/exresult/"+str(eln)+"/o.txt"
            d_path = "/home/hadoop/Projects/exresult/"+str(eln)+"/d.txt"


        # if r_ns == 0:

        #     log = xes_importer.apply(log_path)
        #     tree, nodes = ptml.apply(model_path)
        #     # print(tree)
        #     # pm4py.view_process_tree(tree)

        #     # 原方法
        #     net, im, fm = pt_converter.apply(tree)
        #     # print(net)
        #     o_start_time = time.time()
        #     o_result = at.apply(log, net, im, fm)
        #     o_fitness = call_for_alignment.evaluate_alignment(o_result)
        #     o_end_time = time.time()

        d_pl = [2,4,6,8,10]
        for d_p in d_pl:
            start_r_index = 2
            start_c_index = 3
            if eln == 1 or eln == "prAm6":
                start_r_index = 2
            elif eln == 2 or eln == "prBm6":
                start_r_index = 10
            elif eln == 3 or eln == "prCm6":
                start_r_index =18
            elif eln == 4 or eln == "prEm6":
                start_r_index =26
            elif eln == 5:
                start_r_index = 34
            if d_p == 2:
                start_c_index = 3
            elif d_p == 4:
                start_c_index = 8
            elif d_p == 6:
                start_c_index = 13
            elif d_p == 8:
                start_c_index = 18
            elif d_p == 10:
                start_c_index =23
            if r_ns == 1:
                start_c_index += 1
            elif r_ns == 2:
                start_c_index += 2
            start_c_index+=2
            print(d_p)
            log = xes_importer.apply(log_path)
            tree, nodes = ptml.apply(model_path)

            priority_method = NodePriorityVariants.VERSION_ACTIVITY_NUN
            trace_allocate_method = TraceAllocateVariants.VERSION_TAR_SET
            

            decom_percent = d_p
            conformance_alg = ConformanceVariants.VERSION_ALIGNMENT


            d_start_time = time.time()
            if conformance_alg == ConformanceVariants.VERSION_ALIGNMENT:
                model_cost = version2.get_tree_cost(tree)//10000
                
            d_result = decom_conformance(start_r_index,start_c_index, r_path,s_path, log, tree, nodes, decom_percent, conformance_alg, priority_method, trace_allocate_method)
            c = list()
            c.append(d_result)
            start_r_index += 3
            file = op.load_workbook(r_path)
            sheet = file["Sheet1"]
            if conformance_alg == ConformanceVariants.VERSION_ALIGNMENT:
                d_fitness = version2.get_method(conformance_alg).evaluate(c, model_cost)
                sheet.cell(start_r_index,start_c_index,d_fitness[1]["log_fitness"])
                start_r_index += 1
                sheet.cell(start_r_index,start_c_index,str(d_fitness[1]))
            else:
                d_fitness = version2.get_method(conformance_alg).evaluate(c)
                sheet.cell(start_r_index,start_c_index,d_fitness["fitness"])
                start_r_index += 1
                sheet.cell(start_r_index,start_c_index,str(d_fitness))
            d_end_time = time.time()

            start_r_index += 1
            sheet.cell(start_r_index,start_c_index, d_end_time-d_start_time)
            file.save(s_path)
            if r_ns == 0:
                file = op.load_workbook(r_path)
                sheet = file["Sheet1"]
                allocate_error = 0
                o_r = dict()
                for r in o_result:
                    ci = r['case_id']
                    o_r[ci] =r
                for i in range(len(d_result)):
                    case_id = d_result[i]['case_id']
                    if o_r[case_id]['cost']//10000 != d_result[i]['cost']//10000:
                        allocate_error += 1
                start_r_index += 1
                sheet.cell(start_r_index,start_c_index,allocate_error)
                start_r_index += 1
                sheet.cell(start_r_index,start_c_index,1-float(allocate_error)/float(len(d_result)))
                file.save(s_path)
            


    


