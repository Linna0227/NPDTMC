import gc
import openpyxl as op
# from ctypes.wintypes import PLARGE_INTEGER
from sympy import N, im
from DPMConformanceExtension import Node_Priority
from DPMConformanceExtension import trace_allocate
# from DDPMC import trace_allocate
from DPMConformanceExtension import conformance_algorithm
from pm4py.objects.log.obj import EventLog, Trace, Event
from DPMConformanceExtension.decompose_conformance_method import decomposetree2
import time
import pm4py
from pm4py.util import exec_utils
from pm4py.statistics.variants.log import get as variants_module
from pm4py.algo.conformance.alignments.petri_net.algorithm import Parameters
from copy import copy
from pm4py.util.xes_constants import DEFAULT_NAME_KEY
from pm4py.util.constants import PARAMETER_CONSTANT_ACTIVITY_KEY, PARAMETER_CONSTANT_CASEID_KEY

from DPMConformanceExtension import to_petri_net as pt_converter
from DPMConformanceExtension.obj import Operator

import pm4py.algo.conformance.alignments.petri_net.algorithm as at
from enum import Enum


class NodePriorityVariants(Enum):
    VERSION_ACTIVITY_NUN = Node_Priority.activity_num
    VERSION_STATUS_NUM = Node_Priority.status_num


class TraceAllocateVariants(Enum):
    VERSION_ACTIVITY_SET = trace_allocate.activity_set
    VERSION_TAR_SET = trace_allocate.tar_set
    VERSION_ACTIVITY_SET_IMPROVE = trace_allocate.activity_set_improve
    VERSION_RANDOM_SELECTION = trace_allocate.random_selection


class ConformanceVariants(Enum):
    VERSION_ALIGNMENT = conformance_algorithm.call_for_alignment
    VERSION_TOKEN_REPLAY = conformance_algorithm.call_for_token


DEFAULT_NodePriorityVARIANT = Node_Priority.activity_num
DEFAULT_TraceAllocateVARIANT = TraceAllocateVariants.VERSION_ACTIVITY_SET
DEFAULT_ConformanceVARIANT = ConformanceVariants.VERSION_ALIGNMENT


def get_variants_structure(log, parameters):
    if parameters is None:
        parameters = copy({PARAMETER_CONSTANT_ACTIVITY_KEY: DEFAULT_NAME_KEY})

    parameters = copy(parameters)
    variants_idxs = exec_utils.get_param_value(Parameters.VARIANTS_IDX, parameters, None)
    if variants_idxs is None:
        variants_idxs = variants_module.get_variants_from_log_trace_idx(log, parameters=parameters)

    one_tr_per_var = []
    variants_list = []
    for index_variant, var in enumerate(variants_idxs):
        variants_list.append(var)

    for var in variants_list:
        one_tr_per_var.append(log[variants_idxs[var][0]])

    return variants_idxs, one_tr_per_var


def form_result(log, variants_idxs, all_result):
    al_idx = {}
    for index_variant, variant in enumerate(variants_idxs):
        for trace_idx in variants_idxs[variant]:
            all_result[index_variant]['case_id'] = log[trace_idx].attributes['concept:name']
            al_idx[trace_idx] = all_result[index_variant]

    results_con = []
    for i in range(len(log)):
        results_con.append(al_idx[i])

    return results_con


def get_method(method):
    if isinstance(method, Enum):
        return method.value
    return method


def compute_cost(node):
    if node.operator is None:
        if node.label is not None:
            return 10000
        else:
            return 1
    else:
        if node.operator == Operator.SEQUENCE or node.operator == Operator.PARALLEL:
            cost =0
            for child in node.children:
                cost += compute_cost(child)

        if node.operator == Operator.XOR:
            m = compute_cost(node.children[0])

            for child in node.children[1:]:
                n = compute_cost(child)

                if n < m:
                    m = n
            cost = m
        if node.operator == Operator.LOOP:
            child = node.children[0]
            cost = compute_cost(child)

    return cost


def get_tree_cost(tree):
    root = tree._get_root()
    current_node = root
    tree_cost = compute_cost(current_node)
    return tree_cost


def decom_conformance(start_r_index,start_c_index, r_path, s_path, eventlog, pt, tree_nodes_dict, decom_percent, conformance_algorithm=DEFAULT_ConformanceVARIANT, priority_method=DEFAULT_NodePriorityVARIANT, trace_allocate_method=DEFAULT_TraceAllocateVARIANT):


    new_nodes, xor_to_d, max_layer, d_xor_num = get_method(priority_method).get_priority_of_pt(pt, tree_nodes_dict, decom_percent)

    root = pt._get_root()
    d_s_time = time.time()
    nodes_trees_list = decomposetree2.decompose_tree(pt, new_nodes, root, xor_to_d, max_layer, d_xor_num)
    d_e_time = time.time()
    
    file = op.load_workbook(r_path)
    sheet = file["Sheet1"]
    sheet.cell(start_r_index,start_c_index,d_e_time-d_s_time)
    

    sub_trees_list = nodes_trees_list[root.uid]
    print(len(sub_trees_list))
    start_r_index += 1 
    sheet.cell(start_r_index,start_c_index, len(sub_trees_list))

    get_method(priority_method).clear_v()
    trees = list()

    converter_start = time.time()
    for subtree in sub_trees_list:

        subtree.parent =None

        net, im, fm = pt_converter.apply(subtree)
        trees.append([(net, im, fm), subtree])
    converter_end = time.time()

    start_r_index += 1 
    sheet.cell(start_r_index,start_c_index, converter_end-converter_start)
    file.save(s_path)

    current_log_confor_result = []
    variants_id,variant_traces = get_variants_structure(eventlog, None)
    for trace in variant_traces:

        most_suitable_model = get_method(trace_allocate_method).get_model(trace, trees)

        log=EventLog()
        log.append(trace)
        trace_confor_result = get_method(conformance_algorithm).apply(log, most_suitable_model[0], most_suitable_model[1], most_suitable_model[2])
        trace_confor_result[0]['bwc'] = len(trace)
        current_log_confor_result.append(trace_confor_result[0])
    r_current_log_confor_result=form_result(eventlog,variants_id,current_log_confor_result)

    return r_current_log_confor_result
