from locale import currency
import ptml
import time
import openpyxl as op
from conformance_algorithm import call_for_alignment
from decompose_conformance_method.version2 import NodePriorityVariants
from decompose_conformance_method.version2 import TraceAllocateVariants
from decompose_conformance_method.version2 import ConformanceVariants
from pm4py.objects.log.obj import EventLog, Trace, Event
from decompose_conformance_method import decomposetree2
import time
import pm4py
import to_petri_net as pt_converter
from obj import Operator
from enum import Enum
from decompose_conformance_method import version2

# from matplotlib.pyplot import get
from pm4py.objects.log.obj import EventLog, Trace, Event
from pm4py.util import xes_constants as xes
from pm4py.util import constants as pm4_constants
from pm4py.algo.conformance.tokenreplay import algorithm as token_replay

# import re
from pm4py.algo.conformance.alignments.petri_net import algorithm as alignments
# from pm4py.algo.conformance.alignments.petri_net.algorithm import Variants
# from pytz import NonExistentTimeError
from pyspark import SparkContext
import time
import sys
# from pm4py.objects.conversion.wf_net import converter as wf_net_converter
from pm4py.util import exec_utils
from pm4py.statistics.variants.log import get as variants_module
from pm4py.algo.conformance.alignments.petri_net.algorithm import Parameters
# import hrdecomposetreela as dt
from copy import copy
from pm4py.util.xes_constants import DEFAULT_NAME_KEY
from pm4py.util.constants import PARAMETER_CONSTANT_ACTIVITY_KEY, PARAMETER_CONSTANT_CASEID_KEY

'''describe:
Firstly, the trace and subnets are allocated, 
and fitness of each trace is calculated only with the corresponding subnets.
'''

# compute the cost of model
def get_method(method):
    if isinstance(method, Enum):
        return method.value
    return method


def compute_cost(node):
    if node.operator is None:
        if node.label is not None:
            return 10000
        else:
            return 1
    else:
        if node.operator == Operator.SEQUENCE or node.operator == Operator.PARALLEL:
            cost =0
            for child in node.children:
                cost += compute_cost(child)
                # print(child, ",", cost)
        if node.operator == Operator.XOR:
            m = compute_cost(node.children[0])
            # print("m", m)
            for child in node.children[1:]:
                n = compute_cost(child)
                # print("n", n)
                if n < m:
                    m = n
            cost = m
        if node.operator == Operator.LOOP:
            child = node.children[0]
            cost = compute_cost(child)

    return cost


def get_tree_cost(tree):
    root = tree._get_root()
    current_node = root
    tree_cost = compute_cost(current_node)
    return tree_cost

# covert rdd strings to log 
def covertToXlog(traces):
    log=EventLog()
    parameters={}
    activity_key = parameters[
        pm4_constants.PARAMETER_CONSTANT_ACTIVITY_KEY] if pm4_constants.PARAMETER_CONSTANT_ACTIVITY_KEY in parameters else xes.DEFAULT_NAME_KEY

    for traceString in traces:
        # traceString=re.sub(r'[0-9]+,', '', traceString)
        traceString=traceString.strip('\n')
        eventnames=traceString.split(",")
        trace=Trace()
        trace._set_attributes({'concept:name': eventnames[0]})
        if len(eventnames[1:]) == 0 or eventnames[1:] == ['']:
            log.append(trace)
        else:
            for name in eventnames[1:]:
                event=Event()
                event[activity_key]=name
                trace.append(event)
            log.append(trace)
            

    return log

# call function reformat times: number of partition blocks  (2 times)
def reformat(partitiondata):
    updatedata=list()
    updatedata.append(covertToXlog(partitiondata))
    return iter(updatedata)


def get_variants_structure(log, parameters):
    if parameters is None:
        parameters = copy({PARAMETER_CONSTANT_ACTIVITY_KEY: DEFAULT_NAME_KEY})

    parameters = copy(parameters)
    variants_idxs = exec_utils.get_param_value(Parameters.VARIANTS_IDX, parameters, None)
    if variants_idxs is None:
        variants_idxs = variants_module.get_variants_from_log_trace_idx(log, parameters=parameters)

    one_tr_per_var = []
    variants_list = []
    for index_variant, var in enumerate(variants_idxs):
        variants_list.append(var)

    for var in variants_list:
        one_tr_per_var.append(log[variants_idxs[var][0]])

    return variants_idxs, one_tr_per_var


def form_result(log, variants_idxs, all_result):
    al_idx = {}
    for index_variant, variant in enumerate(variants_idxs):
        for trace_idx in variants_idxs[variant]:
            al_idx[trace_idx] = all_result[index_variant]

    results_con = []
    for i in range(len(log)):
        results_con.append(al_idx[i])

    return results_con



def cost_alignment(partitiondatas):
    m_net=broadcast_net.value
    aligned_traces=list()
    for ilog in partitiondatas:
        variants_id,variant_traces=get_variants_structure(ilog,None)
        # print(variant_traces)
        aligned_ilog=list()
        for one_trace in variant_traces:
            confer_net = get_method(trace_allocate_method).get_model(one_trace, m_net)
            log=EventLog()
            log.append(one_trace)
            aligned_trace=alignments.apply_log(log,confer_net[0],confer_net[1],confer_net[2])

            aligned_trace[0]['bwc']=len(one_trace)
            aligned_ilog.append(aligned_trace[0])
        r_f_ilog=form_result(ilog,variants_id,aligned_ilog)
        aligned_traces.append(r_f_ilog)
    return iter(aligned_traces)

if __name__=="__main__":
    sc = SparkContext( )
    tree_p = sys.argv[1]
    log_path = sys.argv[2]
    Dthreshold = sys.argv[3]
    partition_n = sys.argv[4]
    nofruntimes = sys.argv[5]

    Dthreshold=int(Dthreshold)
    partition_n=int( partition_n)
    nofruntimes = int(nofruntimes)
    if tree_p in ['L1','L2','L3','L4','L5']:

        tree_path = "/home/hadoop/Projects/testdata/data/"+str(tree_p)+"/"+str(tree_p)+".ptml"
        r_path = r"/home/hadoop/Projects/exresult/statusnum_tarset.xlsx"
        s_path = "/home/hadoop/Projects/exresult/statusnum_tarset.xlsx"
    else:
        tree_path = "/home/hadoop/Projects/testdata/data/BPM2013/Multitree/"+str(tree_p)+".ptml"
        r_path = r"/home/hadoop/Projects/exresult/Astatusnum_tarset.xlsx"
        s_path = "/home/hadoop/Projects/exresult/Astatusnum_tarset.xlsx"
    priority_method = NodePriorityVariants.VERSION_STATUS_NUM
    trace_allocate_method = TraceAllocateVariants.VERSION_TAR_SET

    if tree_p == "L1" or tree_p == "prAm6":
        start_r_index = 2
    elif tree_p == "L2" or tree_p == "prBm6":
        start_r_index = 10
    elif tree_p == "L3" or tree_p == "prCm6":
        start_r_index =18
    elif tree_p == "L4" or tree_p == "prEm6":
        start_r_index =26
    elif tree_p == "L5":
        start_r_index = 34

    if Dthreshold == 2:
        start_c_index = 3
    elif Dthreshold == 4:
        start_c_index = 8
    elif Dthreshold == 6:
        start_c_index = 13
    elif Dthreshold == 8:
        start_c_index = 18
    elif Dthreshold == 10:
        start_c_index =23

    if nofruntimes == 2:
        start_c_index += 1
    elif nofruntimes == 3:
        start_c_index += 2

    file = op.load_workbook(r_path)
    sheet = file["Sheet1"]


    conformance_alg = ConformanceVariants.VERSION_ALIGNMENT


    tree, nodes = ptml.apply(tree_path)

    start_time=time.time()
    best_worse_cost=get_tree_cost(tree)
    best_worse_cost=best_worse_cost//10000
    d_start_time=time.time()
    #node  priority
    new_nodes, xor_to_d, max_layer, d_xor_num = get_method(priority_method).get_priority_of_pt(tree, nodes, Dthreshold)
    # tree decompose
    root = tree._get_root()
    nodes_trees_list = decomposetree2.decompose_tree(tree, new_nodes, root, xor_to_d, max_layer, d_xor_num)
    d_end_time=time.time()
    sheet.cell(start_r_index,start_c_index,d_end_time-d_start_time)
    

    # pt-> pn 
    cv_start_time = time.time()
    sub_trees_list = nodes_trees_list[root.uid]
    start_r_index += 1
    sheet.cell(start_r_index,start_c_index,len(sub_trees_list))
    trees = list()
    for subtree in sub_trees_list:
        net, im, fm = pt_converter.apply(subtree)
        trees.append([(net, im, fm), subtree])
    print("-------------------------------------")
    print("-------------------------------------")
    print("-------------------------------------")
    print("num of subtrees")
    print(len(trees))
    cv_end_time = time.time()
    start_r_index += 1
    sheet.cell(start_r_index,start_c_index,cv_end_time-cv_start_time)
    file.save(s_path)

    sc.addPyFile("/home/hadoop/Projects/DDPMC.zip")
    broadcast_net=sc.broadcast(trees)
    l_start_time=time.time()  
    log = sc.textFile(log_path,partition_n)
    sublogs=log.mapPartitions(reformat)
    l_end_time=time.time()

    c_start_time = time.time()
    if conformance_alg == ConformanceVariants.VERSION_ALIGNMENT:
        d_result = sublogs.mapPartitions(cost_alignment).collect()
        d_fitness = version2.get_method(conformance_alg).evaluate(d_result, best_worse_cost)

    else:
        d_result = sublogs.mapPartitions(cost_token).collect()
        d_fitness = version2.get_method(conformance_alg).evaluate(d_result)
    c_end_time = time.time()

    file = op.load_workbook(r_path)
    sheet = file["Sheet1"]
    start_r_index += 1
    sheet.cell(start_r_index,start_c_index,c_end_time-c_start_time)
    start_r_index += 1
    sheet.cell(start_r_index,start_c_index,c_end_time-start_time)

    if conformance_alg == ConformanceVariants.VERSION_ALIGNMENT:
        start_r_index += 1
        sheet.cell(start_r_index,start_c_index,d_fitness[1]["log_fitness"])
        start_r_index += 1
        sheet.cell(start_r_index,start_c_index,str(d_fitness[1]))
        file.save(s_path)
    else:
        start_r_index += 1
        sheet.cell(start_r_index,start_c_index,d_fitness["fitness"])
        start_r_index += 1
        sheet.cell(start_r_index,start_c_index,str(d_fitness))
        file.save(s_path)

    sc.stop()


    









